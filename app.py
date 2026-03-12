from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import os
from werkzeug.utils import secure_filename
from ai_processor import AIProcessor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json
import tempfile
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'h2o-ai-test-generator-secret-key'

# Use temp directory for production (Render)
if os.environ.get('RENDER'):
    app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()
else:
    app.config['UPLOAD_FOLDER'] = 'uploads'
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour

# Valid credentials
VALID_CREDENTIALS = {
    'ncharala': 'Password1!',
    'test': 'test'
}

@app.route('/')
def index():
    if 'logged_in' in session:
        return redirect(url_for('home'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if username in VALID_CREDENTIALS and VALID_CREDENTIALS[username] == password:
            session.permanent = True
            session['logged_in'] = True
            print(f"User '{username}' logged in successfully. Session ID: {session.get('_id', 'No ID')}")
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': 'Invalid credentials'})
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('test_results', None)
    return redirect(url_for('login'))

@app.route('/home')
def home():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template('home.html')

@app.route('/upload', methods=['POST'])
def upload():
    print("\n" + "="*60)
    print("UPLOAD ENDPOINT CALLED!")
    print("="*60)
    
    if 'logged_in' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})
    
    try:
        files = request.files.getlist('files')
        urls = request.form.getlist('urls')
        
        print(f"\n=== Upload Request ===")
        print(f"Files received: {len(files)}")
        for f in files:
            if f and f.filename:
                print(f"  - {f.filename}")
        print(f"URLs received: {urls}")
        
        if not files and not urls:
            return jsonify({'success': False, 'message': 'No files or URLs provided'})
        
        uploaded_files = []
        
        # Save uploaded files
        for file in files:
            if file and file.filename:
                try:
                    filename = secure_filename(file.filename)
                    file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
                    
                    if file_ext not in ['doc', 'docx', 'pdf', 'txt']:
                        continue
                    
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    uploaded_files.append({
                        'path': filepath,
                        'type': file_ext,
                        'name': filename
                    })
                    print(f"Saved file: {filepath}")
                except Exception as fe:
                    print(f"Error saving file {file.filename}: {str(fe)}")
                    continue
        
        # Filter valid URLs
        valid_urls = [u.strip() for u in urls if u and u.strip()]
        print(f"Valid URLs to process: {valid_urls}")
        
        # Process with AI
        processor = AIProcessor()
        test_results = processor.process_documents(uploaded_files, valid_urls)
        
        print(f"\n=== Processing Complete ===")
        print(f"Generated {len(test_results)} test scenarios")
        
        if test_results and len(test_results) > 0:
            # Store results in session
            session['test_results'] = test_results
            session.modified = True
            
            print(f"Stored {len(test_results)} scenarios in session")
            print(f"Session ID: {session.get('_id', 'No ID')}")
            
            return jsonify({'success': True, 'results': test_results})
        else:
            return jsonify({'success': False, 'message': 'No test cases generated'})
            
    except Exception as e:
        print(f"Error processing: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error processing: {str(e)}'})

@app.route('/get_results')
def get_results():
    if 'logged_in' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'})
    
    # Get from session
    results = session.get('test_results', [])
    
    print(f"\n=== Get Results ===")
    print(f"Results found: {len(results)} scenarios")
    
    if results:
        print(f"First scenario: {results[0]['ts_id']} - {results[0]['scenario_desc'][:50]}")
    
    return jsonify({'success': True, 'results': results})

@app.route('/download_excel')
def download_excel():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    
    # Get from session
    results = session.get('test_results', [])
    
    print(f"\n=== Download Excel ===")
    print(f"Results available: {len(results) if results else 0} scenarios")
    
    if not results or len(results) == 0:
        return "No results to download. Please generate test cases first.", 400
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Define headers
    headers = ['TS#', 'Scenario Description', 'TC#', 'Test Case Steps', 
               'Expected Results', 'Test Data', 'Results', 'Defects', 'Comments']
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Populate data
    row = 2
    for scenario in results:
        ts_id = scenario.get('ts_id', '')
        scenario_desc = scenario.get('scenario_desc', '')
        
        test_cases = scenario.get('test_cases', [])
        for tc in test_cases:
            ws.cell(row=row, column=1, value=ts_id)
            ws.cell(row=row, column=2, value=scenario_desc)
            ws.cell(row=row, column=3, value=tc.get('tc_id', ''))
            ws.cell(row=row, column=4, value=tc.get('steps', ''))
            ws.cell(row=row, column=5, value=tc.get('expected', ''))
            ws.cell(row=row, column=6, value=tc.get('test_data', ''))
            ws.cell(row=row, column=7, value=tc.get('results', ''))
            ws.cell(row=row, column=8, value=tc.get('defects', ''))
            ws.cell(row=row, column=9, value=tc.get('comments', ''))
            row += 1
    
    print(f"Generated Excel with {row-2} test cases")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    
    # Generate unique filename with timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f'H2O_AI_TestCases_{timestamp}.xlsx'
    
    # Save file
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb.save(output_path)
    
    print(f"Excel file saved: {output_path}")
    
    return send_file(output_path, as_attachment=True, download_name=filename)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
